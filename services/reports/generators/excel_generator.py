"""
Black Door — Excel Report Generator
Uses openpyxl to create professionally formatted Excel files.
All monetary values are stored as integers (cents/tiyin) and displayed divided by 100.
"""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter


# ─── Style Constants ─────────────────────────────────────────────────────────
HEADER_FONT = Font(name="Arial", bold=True, size=12, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

TITLE_FONT = Font(name="Arial", bold=True, size=16, color="1F4E79")
SUBTITLE_FONT = Font(name="Arial", size=11, color="666666", italic=True)

DATA_FONT = Font(name="Arial", size=10)
DATA_ALIGNMENT = Alignment(vertical="center", wrap_text=True)
NUMBER_ALIGNMENT = Alignment(horizontal="right", vertical="center")

THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

EVEN_ROW_FILL = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")

TOTAL_FONT = Font(name="Arial", bold=True, size=11)
TOTAL_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

USD_FORMAT = '#,##0.00 "$"'
UZS_FORMAT = '#,##0.00 "сўм"'


def _cents_to_display(value: int | None) -> float:
    """Convert integer cents/tiyin to display format."""
    if value is None:
        return 0.0
    return value / 100.0


def _apply_header_style(cell):
    """Apply header styling to a cell."""
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = HEADER_ALIGNMENT
    cell.border = THIN_BORDER


def _apply_data_style(cell, row_idx: int):
    """Apply data styling to a cell with alternating row colors."""
    cell.font = DATA_FONT
    cell.alignment = DATA_ALIGNMENT
    cell.border = THIN_BORDER
    if row_idx % 2 == 0:
        cell.fill = EVEN_ROW_FILL


def _auto_column_width(ws, col_idx: int, min_width: int = 12, max_width: int = 40):
    """Auto-fit column width based on content."""
    letter = get_column_letter(col_idx)
    max_len = min_width
    for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
        for cell in row:
            if cell.value:
                max_len = max(max_len, min(len(str(cell.value)) + 2, max_width))
    ws.column_dimensions[letter].width = max_len


def generate_income_expense_excel(report_data: dict) -> str:
    """Generate income/expense report as Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Daromad va Xarajat"

    # ─── Title ───────────────────────────────────────────────────────────
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = "BLACK DOOR — Daromad va Xarajat Hisoboti"
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A2:G2")
    subtitle_cell = ws["A2"]
    subtitle_cell.value = f"Davr: {report_data.get('period', '')}"
    subtitle_cell.font = SUBTITLE_FONT
    subtitle_cell.alignment = Alignment(horizontal="center")

    # ─── Summary Row ─────────────────────────────────────────────────────
    ws["A4"] = "Jami Daromad (USD):"
    ws["B4"] = _cents_to_display(report_data.get("total_income_usd", 0))
    ws["B4"].number_format = USD_FORMAT
    ws["A4"].font = TOTAL_FONT
    ws["B4"].font = TOTAL_FONT

    ws["C4"] = "Jami Daromad (UZS):"
    ws["D4"] = _cents_to_display(report_data.get("total_income_uzs", 0))
    ws["D4"].number_format = UZS_FORMAT
    ws["C4"].font = TOTAL_FONT
    ws["D4"].font = TOTAL_FONT

    ws["A5"] = "Jami Xarajat (USD):"
    ws["B5"] = _cents_to_display(report_data.get("total_expense_usd", 0))
    ws["B5"].number_format = USD_FORMAT
    ws["A5"].font = TOTAL_FONT
    ws["B5"].font = TOTAL_FONT

    ws["C5"] = "Jami Xarajat (UZS):"
    ws["D5"] = _cents_to_display(report_data.get("total_expense_uzs", 0))
    ws["D5"].number_format = UZS_FORMAT
    ws["C5"].font = TOTAL_FONT
    ws["D5"].font = TOTAL_FONT

    # ─── Headers ─────────────────────────────────────────────────────────
    headers = ["Sana", "Kategoriya", "Turi", "Summa (USD)", "Summa (UZS)", "Izoh"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col_idx, value=header)
        _apply_header_style(cell)

    # ─── Data Rows ───────────────────────────────────────────────────────
    items = report_data.get("items", [])
    for row_idx, item in enumerate(items, 8):
        ws.cell(row=row_idx, column=1, value=item.get("date", ""))
        ws.cell(row=row_idx, column=2, value=item.get("category", ""))
        ws.cell(row=row_idx, column=3, value=item.get("type", ""))

        usd_cell = ws.cell(row=row_idx, column=4, value=_cents_to_display(item.get("amount_usd")))
        usd_cell.number_format = USD_FORMAT
        usd_cell.alignment = NUMBER_ALIGNMENT

        uzs_cell = ws.cell(row=row_idx, column=5, value=_cents_to_display(item.get("amount_uzs")))
        uzs_cell.number_format = UZS_FORMAT
        uzs_cell.alignment = NUMBER_ALIGNMENT

        ws.cell(row=row_idx, column=6, value=item.get("note", ""))

        # Apply styling
        for col in range(1, 7):
            _apply_data_style(ws.cell(row=row_idx, column=col), row_idx)

    # ─── Auto-fit Columns ────────────────────────────────────────────────
    for col in range(1, 7):
        _auto_column_width(ws, col)

    # ─── Footer ──────────────────────────────────────────────────────────
    footer_row = len(items) + 9
    ws.cell(row=footer_row, column=1, value=f"Yaratilgan: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    ws.cell(row=footer_row, column=1).font = Font(size=8, color="999999", italic=True)

    # ─── Save ────────────────────────────────────────────────────────────
    filename = f"income_expense_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(os.getenv("REPORT_FILES_DIR", "/app/generated_reports"), filename)
    wb.save(filepath)
    return filename


def generate_generic_excel(
    title: str,
    headers: list[str],
    rows: list[list],
    usd_columns: list[int] | None = None,
    uzs_columns: list[int] | None = None,
    sheet_name: str = "Hisobot",
) -> str:
    """
    Generate a generic Excel report.
    usd_columns and uzs_columns are 0-indexed column indices that contain monetary values.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    usd_columns = usd_columns or []
    uzs_columns = uzs_columns or []

    # ─── Title ───────────────────────────────────────────────────────────
    num_cols = len(headers)
    end_col = get_column_letter(num_cols)
    ws.merge_cells(f"A1:{end_col}1")
    title_cell = ws["A1"]
    title_cell.value = f"BLACK DOOR — {title}"
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # ─── Headers ─────────────────────────────────────────────────────────
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        _apply_header_style(cell)

    # ─── Data Rows ───────────────────────────────────────────────────────
    for row_idx, row_data in enumerate(rows, 4):
        for col_idx, value in enumerate(row_data):
            cell = ws.cell(row=row_idx, column=col_idx + 1)

            if col_idx in usd_columns:
                cell.value = _cents_to_display(value)
                cell.number_format = USD_FORMAT
                cell.alignment = NUMBER_ALIGNMENT
            elif col_idx in uzs_columns:
                cell.value = _cents_to_display(value)
                cell.number_format = UZS_FORMAT
                cell.alignment = NUMBER_ALIGNMENT
            else:
                cell.value = value

            _apply_data_style(cell, row_idx)

    # ─── Auto-fit Columns ────────────────────────────────────────────────
    for col in range(1, num_cols + 1):
        _auto_column_width(ws, col)

    # ─── Footer ──────────────────────────────────────────────────────────
    footer_row = len(rows) + 5
    ws.cell(row=footer_row, column=1, value=f"Yaratilgan: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    ws.cell(row=footer_row, column=1).font = Font(size=8, color="999999", italic=True)

    # ─── Save ────────────────────────────────────────────────────────────
    filename = f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(os.getenv("REPORT_FILES_DIR", "/app/generated_reports"), filename)
    wb.save(filepath)
    return filename
